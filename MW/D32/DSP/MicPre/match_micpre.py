import openpyxl
import re

# Component references extracted from micpre.pdf OCR
micpre_refs = set()
ocr_text = """R352 C205 R402 R414 Q115 Q116 R351 R353 C204 R389 R345 R390 R391
R346 R403 R411 C190 R348 D32 R347 D31 Q114 R400 R413 R401 U24 D30 R392 Q113
J19 R393 C203 D29 C206 Q112 Q110 R388 R404 C208 C209 R405 R407 R396 R395 R397
R398 R399 R394 R409 Q98 Q100 Q101 Q102 Q103 Q104 Q111 R406 R408 Q99 R365 R366
R369 R370 R373 R374 C192 C195 C194 C191 R350 C197 R367 C198 R368 C199 R371
C200 R372 C201 R375 C193 C196 U25 R349 R354 R355 R356 R358 R376 R377 R378
R379 R380 R357 R381 R382 R383 R384 R385 R386 R387 Q106 Q107 Q108 Q109 R364
R363 R362 R361 R360 R359"""

# Parse all component references (letter prefix + number)
for token in re.findall(r'[A-Z]+\d+', ocr_text):
    micpre_refs.add(token)

# Also add sub-parts (Q115:A means Q115 exists)
print(f"Total unique micpre refs: {len(micpre_refs)}")
sorted_refs = sorted(micpre_refs, key=lambda x: (re.match(r'[A-Z]+', x).group(), int(re.search(r'\d+', x).group())))
print(f"Sorted: {sorted_refs}")

# Read the BOM
wb = openpyxl.load_workbook('/Users/peterwatts/Stonepower Dropbox/Peter Watts/VSCODE/D24 BOM review/D24 rev B BOM/D24 Analog rev B.xlsx', data_only=True)
ws = wb['BOM']

# Parse BOM - header at row 11
print("\n=== MATCHING MIC PRE COMPONENTS TO BOM ===\n")
matched = []
unmatched_refs = set(micpre_refs)

for row in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=False):
    refs_cell = str(row[2].value) if row[2].value else ""
    if not refs_cell:
        continue
    
    # Expand reference ranges like C1-C5 to individual refs
    bom_refs = set()
    for part in refs_cell.split(','):
        part = part.strip()
        range_match = re.match(r'([A-Z]+)(\d+)-\1?(\d+)', part)
        if range_match:
            prefix = range_match.group(1)
            start = int(range_match.group(2))
            end = int(range_match.group(3))
            for n in range(start, end + 1):
                bom_refs.add(f"{prefix}{n}")
        else:
            ref_match = re.match(r'[A-Z]+\d+', part)
            if ref_match:
                bom_refs.add(ref_match.group())
    
    # Check intersection with micpre refs
    overlap = micpre_refs & bom_refs
    if overlap:
        category = str(row[0].value) if row[0].value else ""
        qty_total = row[1].value if row[1].value else 0
        value = str(row[3].value) if row[3].value else ""
        desc = str(row[4].value) if row[4].value else ""
        mfr = str(row[5].value) if row[5].value else ""
        pkg = str(row[6].value) if row[6].value else ""
        notes = str(row[7].value) if row[7].value else ""
        unit_25 = row[8].value if row[8].value else 0
        total_25 = row[9].value if row[9].value else 0
        unit_1000 = row[10].value if row[10].value else 0
        total_1000 = row[11].value if row[11].value else 0
        brand = str(row[12].value) if row[12].value else ""
        lead = str(row[13].value) if row[13].value else ""
        remark = str(row[14].value) if row[14].value else ""
        
        mic_qty = len(overlap)
        mic_cost_25 = mic_qty * float(unit_25) if unit_25 else 0
        mic_cost_1000 = mic_qty * float(unit_1000) if unit_1000 else 0
        
        matched.append({
            'category': category,
            'refs': sorted(overlap, key=lambda x: int(re.search(r'\d+', x).group())),
            'mic_qty': mic_qty,
            'bom_qty': qty_total,
            'value': value,
            'desc': desc,
            'mfr': mfr,
            'pkg': pkg,
            'notes': notes,
            'unit_25': float(unit_25) if unit_25 else 0,
            'unit_1000': float(unit_1000) if unit_1000 else 0,
            'mic_cost_25': mic_cost_25,
            'mic_cost_1000': mic_cost_1000,
            'brand': brand,
            'lead': lead,
            'remark': remark,
        })
        unmatched_refs -= overlap

# Sort by category then value
matched.sort(key=lambda x: (x['category'], x['value']))

total_25 = 0
total_1000 = 0
total_parts = 0

for m in matched:
    total_25 += m['mic_cost_25']
    total_1000 += m['mic_cost_1000']
    total_parts += m['mic_qty']
    print(f"{m['category']:15s} | Qty:{m['mic_qty']:3d} | {','.join(m['refs'][:5])}{'...' if len(m['refs'])>5 else ''} | {m['value']:20s} | {m['desc'][:30]:30s} | Unit@25: ${m['unit_25']:.4f} | Unit@1k: ${m['unit_1000']:.4f} | MicCost@25: ${m['mic_cost_25']:.4f} | MicCost@1k: ${m['mic_cost_1000']:.4f} | Brand: {m['brand']} | {m['pkg']}")

print(f"\n{'='*80}")
print(f"Total mic pre parts: {total_parts}")
print(f"Total mic pre cost @25pcs:   ${total_25:.4f}")
print(f"Total mic pre cost @1000pcs: ${total_1000:.4f}")
print(f"BOM lines matched: {len(matched)}")
print(f"Unmatched refs: {sorted(unmatched_refs) if unmatched_refs else 'None'}")
